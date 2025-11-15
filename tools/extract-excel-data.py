import openpyxl
import pandas as pd
import json
from pathlib import Path

# Paths
excel_path = r"c:\code\decentralized-institutes-of-health\assets\life-expectancy-healthcare-spending-cost-of-drug-development.xlsx"
output_dir = Path(r"c:\code\decentralized-institutes-of-health\brain\data\extracted")
output_dir.mkdir(parents=True, exist_ok=True)

# Load workbook
wb = openpyxl.load_workbook(excel_path, data_only=True)

print(f"Extracting data from {len(wb.sheetnames)} sheets...")
print("=" * 80)

chart_metadata = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Clean sheet name for filename
    safe_name = sheet_name.replace(" ", "-").lower()

    print(f"\nProcessing: {sheet_name}")

    # Extract data using pandas
    data = []
    for row in ws.iter_rows(values_only=True):
        # Only include rows that have at least one non-None value
        if any(cell is not None for cell in row):
            data.append(row)

    if data:
        # Create DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])  # First row as header

        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')

        # Save to CSV
        csv_path = output_dir / f"{safe_name}.csv"
        df.to_csv(csv_path, index=False)
        print(f"  [OK] Saved data to: {csv_path.name} ({len(df)} rows, {len(df.columns)} columns)")

    # Extract chart metadata
    if hasattr(ws, '_charts') and ws._charts:
        chart_metadata[sheet_name] = []

        for idx, chart in enumerate(ws._charts, 1):
            chart_info = {
                'chart_number': idx,
                'type': type(chart).__name__,
                'title': None,
                'series': []
            }

            # Try to get title
            if hasattr(chart, 'title') and chart.title:
                if hasattr(chart.title, 'tx') and chart.title.tx:
                    if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                        # Try to extract title text
                        try:
                            for p in chart.title.tx.rich.p:
                                for r in p.r:
                                    if hasattr(r, 't') and r.t:
                                        chart_info['title'] = r.t
                        except:
                            pass

            # Try to get series information
            if hasattr(chart, 'series'):
                for series_idx, series in enumerate(chart.series, 1):
                    series_info = {
                        'series_number': series_idx,
                        'x_values': None,
                        'y_values': None
                    }

                    # Get series data ranges
                    if hasattr(series, 'xVal') and series.xVal:
                        if hasattr(series.xVal, 'numRef') and series.xVal.numRef:
                            series_info['x_values'] = str(series.xVal.numRef.f)

                    if hasattr(series, 'val') and series.val:
                        if hasattr(series.val, 'numRef') and series.val.numRef:
                            series_info['y_values'] = str(series.val.numRef.f)

                    chart_info['series'].append(series_info)

            chart_metadata[sheet_name].append(chart_info)
            print(f"  [OK] Chart {idx}: {chart_info['type']}")

# Save chart metadata
metadata_path = output_dir / "chart_metadata.json"
with open(metadata_path, 'w') as f:
    json.dump(chart_metadata, f, indent=2)

print("\n" + "=" * 80)
print(f"[OK] Extraction complete!")
print(f"  - Data files: {output_dir}")
print(f"  - Chart metadata: {metadata_path.name}")
print("=" * 80)

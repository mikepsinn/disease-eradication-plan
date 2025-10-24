import openpyxl
import pandas as pd
import json
from pathlib import Path

# Load the workbook
excel_path = r"c:\code\decentralized-institutes-of-health\assets\life-expectancy-healthcare-spending-cost-of-drug-development.xlsx"
wb = openpyxl.load_workbook(excel_path)

print("=" * 80)
print("EXCEL FILE ANALYSIS")
print("=" * 80)
print(f"\nFile: {excel_path}")
print(f"Number of sheets: {len(wb.sheetnames)}")
print("\nSheet names:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

print("\n" + "=" * 80)
print("DETAILED SHEET INFORMATION")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\n{'=' * 60}")
    print(f"SHEET: {sheet_name}")
    print('=' * 60)

    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Dimensions: {max_row} rows x {max_col} columns")

    # Check for charts
    if hasattr(ws, '_charts') and ws._charts:
        print(f"\nNumber of charts: {len(ws._charts)}")
        for idx, chart in enumerate(ws._charts, 1):
            print(f"\n  Chart {idx}:")
            print(f"    Type: {type(chart).__name__}")
            if hasattr(chart, 'title') and chart.title:
                print(f"    Title: {chart.title}")
            if hasattr(chart, 'anchor'):
                print(f"    Anchor: {chart.anchor}")
    else:
        print("\nNo charts found")

    # Show first few rows of data
    print("\nFirst 10 rows of data:")
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, max_row), values_only=True), 1):
        if any(cell is not None for cell in row):  # Only show rows with data
            print(f"  Row {row_idx}: {row}")
            data.append(row)

    # Check for merged cells
    if ws.merged_cells:
        print(f"\nMerged cells: {len(ws.merged_cells.ranges)} ranges")
        for merged_range in list(ws.merged_cells.ranges)[:5]:  # Show first 5
            print(f"  {merged_range}")

print("\n" + "=" * 80)
print("END OF ANALYSIS")
print("=" * 80)
